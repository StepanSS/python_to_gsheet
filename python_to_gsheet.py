'''
errno = 50159747054
SPREADSHEET_URL = 'google.com'
spr_id = 123

print('%x' % errno)

# url = SPREADSHEET_URL % spr_id
# print(url)
'''
import pygsheets
from oauth2client.service_account import ServiceAccountCredentials
import numpy as numpy
import os
import pprint

from openpyxl import load_workbook
import pandas as pd

# pp = pprint.PrettyPrinter()


def createGSheetReturnLink():
    excelRead = ExcelReader()
    excelRead.load_excel_workbook()
    excel_sheets_names = excelRead.get_sheets_names()
    excel_file_name = excelRead.get_excel_name()
    # print(excel_file_name)

    # CONNECT TO GOOGLE
    gsheet = GSheets()
    title = excel_file_name
    gsheet.create(title)

    # ss_metadata = gsheet.get_ss_list_in_folder()
    # ss_title_list = gsheet.get_ss_title_list(ss_metadata)

    # ss = gsheet.open_spreadsheet()

    pass


class GSheets():
    def __init__(self):
        # Scopes
        self.SCOPES = ['https://spreadsheets.google.com/feeds',
                       'https://www.googleapis.com/auth/drive']
        self.SPREADSHEET_ID = '1UAG_6remfTM-muSHW5HBTpBHJYAo4IUpk1f-vq0qzzs'
        self.gc = pygsheets.authorize(service_account_file="creds.json", scopes=self.SCOPES)
        pass

    def open_spreadsheet(self, title=None, ss_id=None):
        if(title is not None):
            self.ss = self.gc.open(title)
            return self.ss
        if(ss_id is not None):
            self.ss = self.gc.open_by_key(ss_id)
            return self.ss
        # by defoult
        self.ss = self.gc.open_by_key(self.SPREADSHEET_ID)
        # self.worksheet = self.ss.worksheet(0)
        return self.ss

    def create(self, title, template=None, folder=None):
        folder = '1BBeCHUwXmQ44EzHsYRhwqVuR0Pdzk3gj'

        # if file exists - open it, else create new
        ss_metadata = self.get_ss_list_in_folder(folder)
        if(title in self.get_ss_title_list(ss_metadata)):
            print("Spreadsheet with this name already exists")
            self.ss = self.open_spreadsheet(title=title)
            return self.ss
        else:
            self.ss = self.gc.create(title, folder=folder)
            # print(self.new_ss))
            return self.ss

    def get_ss_list_in_folder(self, folder_id=None):
        '''Help function - Get all spreadsheets in specific folders'''
        q = '"' + folder_id + '" in parents'
        metadata = self.gc.drive.list(q=q)
        return metadata

    def get_ss_title_list(self, ss_metadata):
        ss_title_list = list(map(lambda x: x['name'], ss_metadata))
        return ss_title_list

    def get_sheet_list(self):
        sh_list = list(map(lambda x: x.title, self.ss.worksheets()))
        return sh_list

    # def get_sheet(self, title):
    #     ''' Get a sheet by title
    #         Return self and create self.sheet
    #     '''
    #     if title in self.get_sheet_list():
    #         self.sheet = self.ss.worksheet_by_title(title)
    #         print(self.sheet)
    #     else:
    #         print('Sheet not exist')
    #     return self

    def add_sheet(self, title):
        '''Create new Sheet'''
        if title not in self.get_sheet_list():
            self.ss.add_worksheet(title)
        else:
            print('already created')
        return self

    def set_dataframe(self, df):
        self.sheet.set_dataframe(df, "A1", copy_index=False, copy_head=False)
        return self

    def setData(self, data=None):
        if self.worksheet:
            self.worksheet.update_value('A2', data)
        return self


class ExcelReader():
    """Class: Read Excel file and pull df"""

    def __init__(self):
        self.excel_file_dir = None
        pass

    def load_excel_workbook(self, excel_file_name='ConstantPositionSize.xlsx'):
        ''' Load excel file
            Load default file - if file name not specified
        '''
        self.excel_file_name = excel_file_name
        excel_file_dir = self.get_excel_dir()
        excel_file_path = os.path.join(excel_file_dir, excel_file_name)
        only_data = True  # If True cell will return data instead formula
        self.wb = load_workbook(excel_file_path, data_only=only_data)
        return self.wb

    def get_excel_dir(self):
        ''' Return default excel file dir if it does not specified'''
        if self.excel_file_dir is None:
            curr_dir = os.path.dirname(os.path.abspath(__file__))
            excel_file_folder = 'excel_templates'
            self.excel_file_dir = os.path.join(curr_dir, excel_file_folder)
        # print(f'file dir: {curr_dir}')
        return self.excel_file_dir

    def set_excel_dir(self, excel_file_dir=None):
        ''' Set specific excel file dir '''
        if excel_file_dir is not None:
            self.excel_file_dir = excel_file_dir
        return

    def get_sheet_df(self, sheetName):
        sheet = self.wb[sheetName]
        row_count = sheet.max_row
        df = pd.DataFrame(sheet.values)
        return df

    def get_sheets_names(self):
        ''' Get sheet names list
            Return list
        '''
        sheets = self.wb.sheetnames
        return sheets

    def get_excel_name(self):
        file_name = os.path.splitext(self.excel_file_name)[0]
        return file_name


def main(excel_workbook_name=None, client_email=None):
    """ This function do:
            Read Excel templates        - @TODO checks if file exist
            Get Excel FileName          - OK
            Store data as Dataframe     - OK
            Connect to Google drive     - OK
            Create G Spreadsheet        - OK
            Copy all data to new GSheet - OK
            Remove Sheet1 from ss       - OK
            Share GSheet and return link- OK
    """
    # GET EXCEL
    excelRead = ExcelReader()  # @TODO: checks if file exist
    if(excel_workbook_name is not None):
        excelRead.load_excel_workbook(excel_workbook_name)
    else:
        excelRead.load_excel_workbook()
    excel_sheets_names = excelRead.get_sheets_names()
    excel_file_name = excelRead.get_excel_name()

    # CONNECT TO GOOGLE
    gsheet = GSheets()
    title = excel_file_name
    ss = gsheet.create(title)
    # ss = gsheet.open_spreadsheet()

    # Iterate all excel sheets pull df and set data to Google SS
    for excel_sheet_name in excel_sheets_names:
        # store each sheet to df
        df = excelRead.get_sheet_df(excel_sheet_name)

        # add sheet to GSheet
        gsheet.add_sheet(excel_sheet_name)

        # get sheet
        sheet = ss.worksheet_by_title(excel_sheet_name)

        # set data to G sheet
        df = df.fillna("")  # Replace "None" cells data with nothing ("")
        sheet.set_dataframe(df, "A1", copy_index=False, copy_head=False)

    # remove extra 'Sheet1'
    if "Sheet1" in gsheet.get_sheet_list():
        ss.del_worksheet(ss.worksheet_by_title('Sheet1'))

    # share spreadsheet and return the link
    if(client_email is not None):
        ss.share(client_email, role='commenter', type='user')
    else:
        ss.share('', role='reader', type='anyone')
    print(ss.url)
    return ss.url


if __name__ == '__main__':
    # createGSheetReturnLink()
    main()
